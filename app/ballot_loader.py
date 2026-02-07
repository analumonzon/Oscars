from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HEADER_MAP = {
    "category": ["category"],
    "nominee": ["nominee", "nomination", "film", "actor", "actress", "title"],
    "points": ["point"],
}


class BallotError(RuntimeError):
    pass


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-")


def _find_index(headers: list[str], keywords: list[str]) -> int:
    for i, header in enumerate(headers):
        for keyword in keywords:
            if keyword in header:
                return i
    return -1


def _finalize(categories: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for category in categories.values():
        if category["points"] is None:
            raise BallotError(f"Missing points for category '{category['name']}'.")
        if not category["nominees"]:
            raise BallotError(f"No nominees found for category '{category['name']}'.")
        result.append(category)
    return sorted(result, key=lambda item: item["name"].lower())


def _load_from_xlsx(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise BallotError("Ballot sheet is empty.")

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    idx_category = _find_index(headers, HEADER_MAP["category"])
    idx_nominee = _find_index(headers, HEADER_MAP["nominee"])
    idx_points = _find_index(headers, HEADER_MAP["points"])

    if idx_category < 0 or idx_nominee < 0 or idx_points < 0:
        raise BallotError(
            "Could not find Category, Nominee, and Points columns in the ballot sheet."
        )

    categories: dict[str, dict[str, Any]] = {}

    for row in rows[1:]:
        category_val = row[idx_category]
        nominee_val = row[idx_nominee]
        points_val = row[idx_points]

        if category_val is None and nominee_val is None:
            continue

        category = str(category_val).strip() if category_val is not None else ""
        nominee = str(nominee_val).strip() if nominee_val is not None else ""
        points = None

        if points_val is not None and str(points_val).strip() != "":
            try:
                points = int(points_val)
            except (TypeError, ValueError):
                raise BallotError(f"Invalid points value: {points_val!r}")

        if not category and nominee:
            raise BallotError(f"Nominee '{nominee}' is missing a category.")
        if not category:
            continue

        key = slugify(category)
        if key not in categories:
            categories[key] = {"key": key, "name": category, "points": None, "nominees": []}

        if points is not None:
            categories[key]["points"] = points

        if nominee:
            categories[key]["nominees"].append(nominee)

    return _finalize(categories)


def _load_from_csv(path: Path) -> list[dict[str, Any]]:
    categories: dict[str, dict[str, Any]] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        rows = [row for row in reader if any(cell.strip() for cell in row)]

    if not rows:
        raise BallotError("Ballot CSV is empty.")

    header = [cell.strip().lower() for cell in rows[0]]
    has_header = any(keyword in " ".join(header) for keyword in ("category", "nominee", "points"))
    data_rows = rows[1:] if has_header else rows

    for row in data_rows:
        if len(row) < 3:
            raise BallotError("Each CSV row must include category, points, and nominees.")

        category = row[0].strip()
        if not category:
            continue

        try:
            points = int(row[1].strip())
        except ValueError as exc:
            raise BallotError(f"Invalid points value: {row[1]!r}") from exc

        nominees = [cell.strip() for cell in row[2:] if cell.strip()]
        if not nominees:
            raise BallotError(f"No nominees found for category '{category}'.")

        key = slugify(category)
        if key not in categories:
            categories[key] = {"key": key, "name": category, "points": points, "nominees": []}
        elif categories[key]["points"] != points:
            raise BallotError(f"Conflicting points for category '{category}'.")

        categories[key]["nominees"].extend(nominees)

    return _finalize(categories)


def load_ballot(path: str | Path) -> list[dict[str, Any]]:
    ballot_path = Path(path)
    if ballot_path.suffix.lower() == ".csv":
        return _load_from_csv(ballot_path)
    return _load_from_xlsx(ballot_path)
